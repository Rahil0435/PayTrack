from django.shortcuts import render, redirect, get_object_or_404
from django.template import loader
from django.http import HttpResponse, JsonResponse
from .models import Product2, Production2, ProductionHistory2, Invoice2, InvoiceItem2, reg2, login2,Factorysale2,Customer2,PaymentRecord2
from .forms import ProductionForm2, ProductForm2, InvoiceForm2,FactorySaleForm2,Customer2Form,MoneyUpdateForm
from django.db import transaction,IntegrityError
from django.contrib import messages
from django.core.paginator import Paginator
from django.contrib.auth import logout
from django.template.loader import get_template
from xhtml2pdf import pisa
from django.template.loader import render_to_string
from django.db.models import Max
from datetime import datetime
import re,logging
from django.db import connection
from django.http import HttpResponseNotAllowed
from django.views.decorators.cache import never_cache
from django.utils.timezone import now
from django.db.models import Sum,F
from decimal import Decimal,InvalidOperation
from datetime import datetime, date
from django.utils import timezone




# Create your views here.
def add_product2(request):
    if request.method == 'POST':
        form = ProductForm2(request.POST)
        if form.is_valid():
            form.save()
            return redirect('productlist2')
    else:
        form = ProductForm2()
    return render(request, 'add product2.html', {'form': form})

def add_production2(request):
    if request.method == 'POST':
        form = ProductionForm2(request.POST)
        if form.is_valid():
            production = form.save(commit=False)  # Save the form but don't commit yet
            product = production.product  # Get the associated product

            # Debugging output: Before update
            print(f"Before update: Product {product.name} quantity: {product.quantity}")

            # Ensure product quantity is only updated for a new production record
            if not production.pk:  # Check if this is a new record (unsaved)
                product.quantity + production.quantity_added  # Update stock quantity
                product.save()  # Save the product changes

                # Debugging output: After update
                print(f"After update: Product {product.name} quantity: {product.quantity}")
            else:
                print(f"No stock update needed. Production record already exists.")

            production.save()  # Save the production record

            # Debugging output: Production saved
            print(f"Production saved: ID={production.id}, Product={product.name}, Quantity Added={production.quantity_added}")

            # Add production history
            ProductionHistory2.objects.create(
                product=product,
                quantity_produced=production.quantity_added,
                date=production.date
            )

            # Debugging output: Production history saved
            print(f"Production history saved for product {product.name} with quantity {production.quantity_added}")

            return redirect('productlist2')  # Redirect to product list after successful save
    else:
        form = ProductionForm2()

    return render(request, 'addproduction2.html', {'form': form})

def product_list2(request):
    products = Product2.objects.all().order_by('price')  # Sorting by price (ascending)
    return render(request, 'Stock_list2.html', {'products': products})

def production_history2(request):
    history = ProductionHistory2.objects.all()
    return render(request, 'production_history2.html', {'history': history})

def get_stock2(request, product_id=None):
    try:
        if product_id:
            product = Product2.objects.get(id=product_id)
            return JsonResponse({'product_name': product.name, 'stock': product.quantity})
        
        # ✅ Sort products by price (ascending order)
        products = Product2.objects.all().order_by('price').values('id', 'name', 'price', 'quantity')
        return JsonResponse({'products': list(products)})

    except Product2.DoesNotExist:
        return JsonResponse({'error': 'Product not found'}, status=404)
    

def createinvoice2(request):
    invoice_form = InvoiceForm2(request.POST or None)

    if request.method == 'POST':
        if invoice_form.is_valid():
            try:
                with transaction.atomic():
                    # Generate invoice number
                    last_invoice = Invoice2.objects.order_by('-id').first()
                    if last_invoice and last_invoice.invoice_number:
                        match = re.search(r'INV2-(\d+)', last_invoice.invoice_number)
                        new_number = int(match.group(1)) + 1 if match else 1
                    else:
                        new_number = 1
                    generated_invoice_number = f"INV2-{new_number:05d}"

                    # Save form data
                    invoice = invoice_form.save(commit=False)
                    invoice.invoice_number = generated_invoice_number
                    invoice.customer = request.POST.get("customer")

                    location_id = request.POST.get("location")
                    customer_obj = None
                    if location_id:
                        try:
                            customer_obj = Customer2.objects.get(id=location_id)
                            invoice.location = customer_obj
                        except Customer2.DoesNotExist:
                            messages.error(request, "Selected location is invalid.")
                            return redirect('createinvoice2')

                    invoice.save()

                    # Get form product and accessory inputs
                    products = request.POST.getlist('products[]')
                    quantities = request.POST.getlist('quantities[]')
                    accessory_quantities = request.POST.getlist('accessory_quantity')
                    accessory_prices = request.POST.getlist('accessory_price')
                    e_way = request.POST.get('e_way', "0")
                    sp_discount = request.POST.get("sp_discount", "0")

                    # Parse decimal inputs
                    try:
                        e_way = Decimal(str(e_way).strip()) if e_way.strip() else Decimal(0)
                    except ValueError:
                        messages.error(request, "Invalid transportation charge entered.")
                        invoice.delete()
                        return redirect('createinvoice2')

                    try:
                        sp_discount = Decimal(str(sp_discount).strip()) if sp_discount.strip() else Decimal(0)
                    except InvalidOperation:
                        messages.error(request, "Invalid special discount entered.")
                        invoice.delete()
                        return redirect('createinvoice2')

                    if not any(products) and not any(accessory_prices):
                        messages.error(request, "Please add at least one product or accessory.")
                        invoice.delete()
                        return redirect('createinvoice2')

                    product_total = Decimal(0)
                    accessory_total = Decimal(0)
                    valid_items = 0

                    # Handle products
                    for i in range(len(products)):
                        try:
                            product_id = products[i].strip()
                            quantity = int(quantities[i].strip()) if quantities[i].strip() else 0

                            if product_id and quantity > 0:
                                product = get_object_or_404(Product2, id=product_id)
                                if product.quantity >= quantity:
                                    price = Decimal(str(product.price))
                                    subtotal = price * quantity
                                    product_total += subtotal

                                    InvoiceItem2.objects.create(
                                        invoice=invoice,
                                        product=product,
                                        quantity=quantity,
                                        price=price,
                                        subtotal=subtotal
                                    )
                                    product.quantity -= quantity
                                    product.save()
                                    valid_items += 1
                                else:
                                    messages.warning(request, f"Not enough stock for {product.name}. Skipped.")
                        except (ValueError, Product2.DoesNotExist) as e:
                            messages.error(request, f"Invalid product selection: {e}")
                            invoice.delete()
                            return redirect('createinvoice2')

                    # Handle accessories
                    for j in range(len(accessory_prices)):
                        try:
                            accessory_price = Decimal(str(accessory_prices[j]).strip()) if accessory_prices[j].strip() else Decimal(0)
                            accessory_quantity = int(accessory_quantities[j].strip()) if accessory_quantities[j].strip() else 0

                            if accessory_price > 0 and accessory_quantity > 0:
                                subtotal = accessory_price
                                accessory_total += subtotal

                                InvoiceItem2.objects.create(
                                    invoice=invoice,
                                    product=None,
                                    quantity=accessory_quantity,
                                    price=accessory_price,
                                    subtotal=subtotal
                                )
                                valid_items += 1
                        except ValueError as e:
                            messages.error(request, f"Invalid accessory price or quantity entered: {e}")
                            invoice.delete()
                            return redirect('createinvoice2')

                    if valid_items == 0:
                        invoice.delete()
                        messages.error(request, "No valid items to create an invoice.")
                        return redirect('createinvoice2')

                    # Apply product discount
                    discount_percentage = Decimal(str(invoice.discount_percentage or 0))
                    discount_amount = (product_total * discount_percentage) / Decimal(100)
                    final_product_total = product_total - discount_amount

                    # Calculate full total
                    total_invoice_amount = final_product_total + accessory_total + e_way - sp_discount

                    # Set original total (DO NOT reduce by advance)
                    invoice.total_amount = total_invoice_amount

                    # Deduct advance from money_got only
                    advance_used = Decimal(0)
                    if customer_obj and customer_obj.advance_amount > 0:
                        advance_available = customer_obj.advance_amount
                        if advance_available >= total_invoice_amount:
                            advance_used = total_invoice_amount
                            customer_obj.advance_amount -= total_invoice_amount
                        else:
                            advance_used = advance_available
                            customer_obj.advance_amount = Decimal('0.00')
                        customer_obj.save()

                    invoice.money_got = advance_used
                    invoice.balance_amount = total_invoice_amount - advance_used
                    invoice.e_way = e_way
                    invoice.advance_used = advance_used
                    invoice.save()

                    messages.success(request, f"Invoice {invoice.invoice_number} created successfully!")
                    return redirect('invoicelist2')

            except IntegrityError as e:
                messages.error(request, f"Invoice creation failed due to a system error: {e}")
            except Exception as e:
                messages.error(request, f"An unexpected error occurred: {e}")

    products = Product2.objects.all().order_by('price')
    location = Customer2.objects.all()
    return render(request, 'create invoice2.html', {
        'invoice_form': invoice_form,
        'products': products,
        'locations': location,
    })



def invoice_list2(request):
    invoices = Invoice2.objects.all().order_by('-date','-id')  # Order by date (newest first)
    return render(request, 'invoice list2.html', {'invoices': invoices})

def invoice_detail2(request, invoice_id):
    invoice = get_object_or_404(Invoice2, id=invoice_id)
    items = InvoiceItem2.objects.filter(invoice=invoice).order_by('price')
    total_qty = sum(item.quantity for item in items if item.product) 
    subtotal = sum(item.quantity * item.price for item in items)
    discount_amount = (subtotal * invoice.discount_percentage) / 100 if invoice.discount_percentage else 0
    accessory=Invoice2.accessory_price
    accessory = Decimal(getattr(invoice, "accessory_price", 0)) 
    e_way=Invoice2.e_way
    e_way = Decimal(getattr(invoice, "e_way", 0)) 
    sp_discount = Invoice2.sp_discount
    sp_discount = Decimal(getattr(invoice, "sp_discount", 0))
    total_amount = subtotal - discount_amount + accessory + e_way
    context = {
        'invoice': invoice,
        'items': items,
        'subtotal': subtotal,
        'discount_amount': discount_amount,
        'total_amount': total_amount,
        'total_qty': total_qty,
        'accessory':accessory,
        'e_way':e_way,
        'sp_discount':sp_discount,
    }
    return render(request, 'invoice details2.html', context)

def Login2(request):
    if request.method == 'POST':
        uname = request.POST.get("uname")
        psw = request.POST.get("pswd")
        try:
            l = login2.objects.get(uname=uname, pswd=psw)
            if l.utype == "user":
                request.session["uid"] = l.uid_id
                return HttpResponse("<script>alert('Welcome User');window.location='/userhome2';</script>")
            elif l.utype == "admin":
                request.session["uid"] = l.uid_id
                return HttpResponse("<script>alert('Welcome Admin');window.location='/adminhome2';</script>")

        except login2.DoesNotExist:
            return HttpResponse("<script>alert('Invalid username or password.');window.location='/login2';</script>")

    template = loader.get_template("login2.html")
    context = {}
    return HttpResponse(template.render(context, request))

def registration2(request):
    if request.method == "POST":
        r = reg2()
        r.name = request.POST.get("Name")
        r.phno = request.POST.get("phno")
        r.save()
        id = reg2.objects.latest("id").id
        l = login2()
        l.uname = request.POST.get("uname")
        l.pswd = request.POST.get("psw")
        l.utype = 'user'
        l.uid_id = id
        l.save()
        return HttpResponse("<script>alert('Registered successfully');window.location='/login2';</script>")

    template = loader.get_template("reg2.html")
    context = {}
    return HttpResponse(template.render(context, request))

@never_cache
def userhome2(request):
    template = loader.get_template("userhome2.html")
    context = {}
    return HttpResponse(template.render(context, request))

def logoutview2(request):
    logout(request)  # Logs out the user
    request.session.flush()  # Clears session data

    response = redirect('login2')  # Redirects to login page
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'

    return response

def generate_pdf2(request, invoice_id):
    # Fetch invoice data from your database
    invoice = get_object_or_404(Invoice2, pk=invoice_id)  # Adjust according to your model
    items = invoice.items.all()  # Assuming items are related to the invoice

    # Render the HTML template with the invoice data
    html_content = render_to_string('invoice template.html', {'invoice': invoice, 'items': items})
    
    # Create the response and set the appropriate content type for PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="invoice.pdf"'
    
    # Generate the PDF from the HTML string
    pisa_status = pisa.CreatePDF(html_content, dest=response)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="invoice.pdf"'
    
    # Generate the PDF from the HTML string
    pisa_status = pisa.CreatePDF(html_content, dest=response)
    
    # If there are errors, return an error message
    if pisa_status.err:
        return HttpResponse("Error generating PDF", status=500)

    return response



def delete_invoice2(request, invoice_id):
    if request.method == 'POST':  # Only allow POST
        invoice = get_object_or_404(Invoice2, id=invoice_id)

        invoice_items = InvoiceItem2.objects.filter(invoice=invoice)
        for item in invoice_items:
            if item.product: 
                item.product.quantity += item.quantity
                item.product.save() 

        invoice.delete()  # Now delete the invoice
        messages.success(request, f"Invoice {invoice.invoice_number} deleted successfully! Product stock has been updated.")
        return HttpResponse("<script>alert('Invoice deleted successfully! Product stock has been updated.');window.location='/invoicelist2';</script>")

    return HttpResponseNotAllowed(['POST']) 

def adminhome2(request):
    template = loader.get_template("adminhome2.html")
    context = {}
    return HttpResponse(template.render(context, request))

def sales_report2(request):
    invoices = Invoice2.objects.all()
    factory_sales = Factorysale2.objects.all()
    today = now().date()
    start_date = request.GET.get('start_date', today)
    end_date = request.GET.get('end_date', today)

    # Apply filtering if both dates are provided
    if start_date and end_date:
        invoices = invoices.filter(date__range=[start_date, end_date])
        factory_sales = factory_sales.filter(date__range=[start_date, end_date])

    # Calculate total sales from both Invoice and Factorysale tables
    total_invoice_sales = sum(invoice.total_amount for invoice in invoices)
    total_factory_sales = sum(factory_sale.total_amount for factory_sale in factory_sales)

    # Overall total sales
    total_sales = total_invoice_sales + total_factory_sales

    context = {
        'invoices': invoices,
        'factory_sales': factory_sales,
        'total_invoice_sales': total_invoice_sales,
        'total_factory_sales': total_factory_sales,
        'total_sales': total_sales,
        'start_date': start_date,
        'end_date': end_date,
    }
    return render(request, 'sales_report2.html', context)

def product_sales_report2(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if not start_date:
        start_date = datetime.today().date()
    else:
        try:
            start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
        except ValueError:
            return render(request, "product_sales_report2.html", {"error": "Invalid start date format"})

    if not end_date:
        end_date = datetime.today().date()
    else:
        try:
            end_date = datetime.strptime(end_date, "%Y-%m-%d").date()
        except ValueError:
            return render(request, "product_sales_report2.html", {"error": "Invalid end date format"})
    sales_data = (
        InvoiceItem2.objects
        .filter(invoice__date__range=[start_date, end_date]) 
        .values('product__name') 
        .annotate(
            total_quantity=Sum('quantity'), 
        )
        .order_by('-total_quantity')
    )
    return render(request, "product_sales_report2.html", {
        "sales_data": sales_data,
        "start_date": start_date,
        "end_date": end_date
    })

def stock_edit2(request):
    """View to display and edit stock list by the admin"""
    products = Product2.objects.all()

    if request.method == "POST":
        for product in products:
            field_name = f'quantity_{product.id}'
            if field_name in request.POST:
                new_quantity = request.POST[field_name]
                try:
                    new_quantity = int(new_quantity)
                    if new_quantity >= 0:
                        product.quantity = new_quantity
                        product.save()
                    else:
                        messages.error(request, f"Invalid quantity for {product.name}")
                except ValueError:
                    messages.error(request, f"Invalid input for {product.name}")

        messages.success(request, "Stock updated successfully!")
        return redirect('stock_edit2')  # Redirect to avoid resubmission

    return render(request, 'stock_edit2.html', {'products': products})

import json
from decimal import Decimal

def decimal_to_float2(obj):
    if isinstance(obj, Decimal):
        return float(obj)
    raise TypeError

def edit_invoice2(request, invoice_id):
    invoice = get_object_or_404(Invoice2, id=invoice_id)

    if request.method == "POST":
        try:
            customer = request.POST.get("customer", invoice.customer)
            date = request.POST.get("date", invoice.date)
            discount_percentage = request.POST.get("discount_percentage", invoice.discount_percentage)
            accessory_price = request.POST.get("accessory_price", invoice.accessory_price)
            e_way = request.POST.get("e_way", invoice.e_way)

            # Convert to Decimal
            discount_percentage = Decimal(discount_percentage) if discount_percentage else Decimal(0)
            accessory_price = Decimal(accessory_price) if accessory_price else Decimal(0)
            e_way = Decimal(e_way) if e_way else Decimal(0)

            product_ids = request.POST.getlist("products[]")
            quantities = request.POST.getlist("quantities[]")

            if not product_ids or not quantities:
                return JsonResponse({"success": False, "message": "Products and quantities are required."})

            if len(product_ids) != len(quantities):
                return JsonResponse({"success": False, "message": "Mismatched product and quantity count."})

            # ✅ Restore stock before deleting old invoice items
            for item in invoice.invoice_items.all():
                if item.product:
                    item.product.quantity += item.quantity
                    item.product.save()

            # ✅ Clear existing invoice items
            invoice.invoice_items.all().delete()

            # ✅ Process new items
            product_total = Decimal(0)
            skipped_items = []
            for product_id, quantity in zip(product_ids, quantities):
                product = Product2.objects.filter(id=product_id).first()

                if not product:
                    skipped_items.append(f"Product ID {product_id} not found")
                    continue

                try:
                    quantity = int(quantity)
                except ValueError:
                    skipped_items.append(f"Invalid quantity for {product.name}")
                    continue

                if product.quantity < quantity:
                    skipped_items.append(f"Not enough stock for {product.name}")
                    continue

                price = product.price
                subtotal = price * quantity
                product_total += subtotal

                # ✅ Deduct the new quantity from stock
                product.quantity -= quantity
                product.save()

                # ✅ Create new invoice item
                InvoiceItem2.objects.create(
                    invoice=invoice,
                    product=product,
                    quantity=quantity,
                    price=price,
                    subtotal=subtotal,
                )

            # ✅ Apply discount only to product total
            discount_amount = (product_total * discount_percentage) / Decimal(100)
            final_product_total = product_total - discount_amount

            # ✅ Update invoice details
            invoice.customer = customer
            invoice.date = date
            invoice.discount_percentage = discount_percentage
            invoice.accessory_price = accessory_price
            invoice.e_way = e_way

            # ✅ Final total = discounted product total + accessory + e-way - special discount
            final_amount = final_product_total + accessory_price + e_way - invoice.sp_discount
            invoice.total_amount = final_amount

            invoice.save()

            if skipped_items:
                skipped_message = f"Invoice updated, but some items were skipped due to issues: {', '.join(skipped_items)}"
                return HttpResponse(f"<script>alert('{skipped_message}');window.location='/invoicelist2';</script>")
            else:
                return HttpResponse("<script>alert('Invoice Updated successfully!');window.location='/invoicelist2';</script>")

        except ValueError:
            return JsonResponse({"success": False, "message": "Invalid number format."})
        except Exception as e:
            return JsonResponse({"success": False, "message": str(e)})

    # ✅ Load data for rendering
    invoice_items = InvoiceItem2.objects.filter(invoice=invoice, product__isnull=False)
    products = Product2.objects.exclude(name__icontains='Accessory')

    invoice_data = {
        "id": invoice.id,
        "customer": str(invoice.customer),
        "date": invoice.date.strftime("%Y-%m-%d"),
        "discount": float(invoice.discount_percentage) if invoice.discount_percentage else 0.0,
        "accessory_price": float(invoice.accessory_price),
        "e_way": float(invoice.e_way),
        "total": float(invoice.total_amount) if invoice.total_amount else 0.0,
        "items": [
            {
                "product_id": item.product.id,
                "quantity": item.quantity,
                "price": float(item.price),
                "subtotal": float(item.subtotal),
            } for item in invoice_items
        ]
    }

    return render(request, 'Edit invoice2.html', {
        'invoice': invoice,
        'invoice_json': json.dumps(invoice_data),
        'products': products
    })


def create_factory_sale2(request):
    if request.method == 'POST':
        form = FactorySaleForm2(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Factory Sale added successfully!")
            return redirect('factorysalelist2')  
        else:
            messages.error(request, "Error in form submission.")

    else:
        form = FactorySaleForm2()

    return render(request, 'create_factory_sale2.html', {'form': form})

def factory_sale_list2(request):
    sales = Factorysale2.objects.all().order_by('-date')
    return render(request, 'factory_sale_list2.html', {'sales': sales})

def delete_factory_sale2(request, sale_id):
    sale = get_object_or_404(Factorysale2, id=sale_id)
    sale.delete()
    messages.success(request, "Factory Sale deleted successfully!")
    return redirect('factorysalelist2')

def addcustomer2(request):
    if request.method == 'POST':
        form = Customer2Form(request.POST)
        if form.is_valid():
            form.save()
            return redirect('viewcustomer2') 
    else:
        form = Customer2Form()
    return render(request, 'Add_customer.html', {'form': form})

def customer2_list(request):
    customers = Customer2.objects.all().order_by('name')  # Optional sorting
    total_advance = customers.aggregate(total=Sum('advance_amount'))['total'] or Decimal('0.00')

    return render(request, 'customer_list.html', {
        'customers': customers,
        'total_advance': total_advance,
    })



from decimal import Decimal
from datetime import datetime, date
from django.db.models import Sum
from django.shortcuts import render, redirect
from django.utils import timezone

def location_report2(request):
    location_id = request.GET.get('location')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Set default dates to today's date if not provided
    today = date.today()
    if not start_date:
        start_date = today.strftime('%Y-%m-%d')
    if not end_date:
        end_date = today.strftime('%Y-%m-%d')

    # Base invoice queryset
    invoices = Invoice2.objects.all().order_by('date', '-id')

    # Filter by location if selected
    if location_id:
        invoices = invoices.filter(location__id=location_id)

    # Filter by date range
    try:
        parsed_start = datetime.strptime(start_date, "%Y-%m-%d").date()
        parsed_end = datetime.strptime(end_date, "%Y-%m-%d").date()
        invoices = invoices.filter(date__gte=parsed_start, date__lte=parsed_end)
    except ValueError:
        pass  # Optional: handle date parsing errors

    # Sort by date DESC, then ID DESC
    invoices = invoices.order_by('-date', '-id')

    # Calculate totals
    total_sales = invoices.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
    raw_total_balance = invoices.aggregate(total=Sum('balance_amount'))['total'] or Decimal('0')
    if raw_total_balance < 0:
        display_total_balance = f"+{abs(raw_total_balance)}"
    else:
        display_total_balance = str(raw_total_balance)
    total_got = invoices.aggregate(total=Sum('money_got'))['total'] or Decimal('0')

    # Calculate total advance amount for customers in selected location
    customers = Customer2.objects.all()
    if location_id:
        customers = customers.filter(id=location_id)

    total_advance = customers.aggregate(total=Sum('advance_amount'))['total'] or 0
    # Handle POST (update money_got)
    if request.method == 'POST':
        count = int(request.POST.get('invoice_count', 0))
        for i in range(1, count + 1):
            invoice_id = request.POST.get(f'invoice_id_{i}')
            money_got = request.POST.get(f'money_got_{i}')
            if invoice_id:
                try:
                    invoice = Invoice2.objects.get(id=invoice_id)
                    added_amount = Decimal(money_got or '0')
                    if added_amount > 0:
                        if invoice.original_money_got is None:
                            invoice.original_money_got = invoice.money_got
                        invoice.money_got += added_amount
                        invoice.balance_amount = invoice.total_amount - invoice.money_got
                        invoice.save()

                        # Record the payment
                        PaymentRecord2.objects.create(
                            invoice=invoice,
                            amount=added_amount,
                            date=timezone.now().date(),
                        )
                except (Invoice2.DoesNotExist, ValueError, TypeError) as e:
                    print(f"Error processing invoice {invoice_id}: {e}")
                    continue

        # Redirect to refresh the page with current filters
        return redirect(
            f"{request.path}?location={location_id}&start_date={start_date}&end_date={end_date}"
            if location_id else request.path
        )

    # Add formatted balance for display
    for invoice in invoices:
        if invoice.balance_amount < 0:
            invoice.display_balance_amount = f"+{abs(invoice.balance_amount)}"
        else:
            invoice.display_balance_amount = str(invoice.balance_amount)

    # Render the template with total advance included
    context = {
        'locations': Customer2.objects.all(),
        'selected_location_id': int(location_id) if location_id else '',
        'selected_start_date': start_date,
        'selected_end_date': end_date,
        'invoices': invoices,
        'total_sales': total_sales,
        'total_balance': raw_total_balance,
        'display_total_balance': display_total_balance,
        'total_got': total_got,
        'total_advance': total_advance,  # Pass total advance here
    }
    return render(request, 'location_report.html', context)




import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from .models import Invoice2

def export_invoices_to_excel2(request):
    location_id = request.GET.get('location')
    invoices = Invoice2.objects.all()

    if location_id:
        invoices = invoices.filter(location__id=location_id)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Invoices'

    headers = [
        'Invoice No', 'Date', 'Customer', 'Location',
        'Total Amount', 'Money Got', 'Balance Amount'
    ]
    sheet.append(headers)

    for invoice in invoices:
        sheet.append([
            invoice.invoice_number,
            invoice.date.strftime('%Y-%m-%d'),
            str(invoice.customer),
            str(invoice.location),
            float(invoice.total_amount),
            float(invoice.money_got),
            float(invoice.balance_amount),
        ])

    # Auto-adjust column widths
    for column_cells in sheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 3

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=location_wise_invoices.xlsx'
    workbook.save(response)
    return response

from django.utils import timezone

from django.db.models import Sum
from .models import PaymentRecord2, Invoice2

def transaction_report2(request):
    # Get the filter parameters from the request
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    selected_location = request.GET.get('location', '')

    # Get distinct locations from Invoice2
    locations = Customer2.objects.filter(invoice2__isnull=False).distinct()

    # Start with all payments
    payments = PaymentRecord2.objects.select_related('invoice')

    # Filter by date
    if start_date:
        payments = payments.filter(date__gte=start_date)
    if end_date:
        payments = payments.filter(date__lte=end_date)

    # Filter by selected location
    if selected_location:
        payments = payments.filter(invoice__location__id=selected_location)

    # Calculate total amount paid
    total_amount = payments.aggregate(Sum('amount'))['amount__sum'] or 0

    context = {
        'payments': payments,
        'start_date': start_date,
        'end_date': end_date,
        'locations': locations,
        'selected_location': locations,
        'total_amount': total_amount,
    }

    return render(request, 'transaction_report.html', context)


def delete_transaction2(request, payment_id):
    # Get the payment record or return 404 if not found
    payment = get_object_or_404(PaymentRecord2, pk=payment_id)

    # Delete the payment record
    payment.delete()

    # Get the start_date and end_date from GET parameters to preserve the filter
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Build the redirect URL, including the filter parameters if present
    redirect_url = 'transaction_report'
    if start_date and end_date:
        redirect_url += f"?start_date={start_date}&end_date={end_date}"

    # Redirect to the transaction report page
    return redirect(redirect_url)

# def undo_invoice_money_got(request, invoice_id):
#     invoice = get_object_or_404(Invoice2, id=invoice_id)
#     if request.method == "POST" and invoice.original_money_got is not None:
#         invoice.money_got = invoice.original_money_got
#         invoice.balance_amount = invoice.total_amount - invoice.money_got
#         invoice.original_money_got = None  # reset it after undo
#         invoice.save()
#     return redirect(request.META.get('HTTP_REFERER', 'location_report2'))

def delete_customer2(request, customer_id):
    customer = get_object_or_404(Customer2, id=customer_id)
    customer.delete()
    return redirect('viewcustomer2')


def update_advance_amount2(request, customer_id):
    if request.method == 'POST':
        customer = get_object_or_404(Customer2, id=customer_id)
        try:
            customer.advance_amount = int(request.POST.get('advance_amount') or 0)
            customer.save()
        except ValueError:
            pass  # Handle invalid input gracefully
    return redirect('viewcustomer2') 

def advance_usage_report2(request):
    invoices_with_advance = Invoice2.objects.filter(advance_used__gt=0)

    total_advance_used = invoices_with_advance.aggregate(total=Sum('advance_used'))['total'] or 0

    return render(request, 'advance_usage_report2.html', {
        'invoices': invoices_with_advance,
        'total_advance_used': total_advance_used,
    })