from django.shortcuts import render, redirect, get_object_or_404
from django.template import loader
from django.http import HttpResponse, JsonResponse
from .models import Product, Production, ProductionHistory, Invoice, InvoiceItem, reg, login,Factorysale,Customer,PaymentRecord
from .forms import ProductionForm, ProductForm, InvoiceForm,FactorySaleForm,MoneyUpdateForm2,CustomerForm
from django.db import transaction,IntegrityError
from django.contrib import messages
from django.core.paginator import Paginator
from django.contrib.auth import logout
from django.template.loader import get_template
from xhtml2pdf import pisa
from django.template.loader import render_to_string
from django.db.models import Max
from datetime import datetime
import re
from django.db import connection
from django.http import HttpResponseNotAllowed
from django.views.decorators.cache import never_cache
from django.utils.timezone import now
from django.db.models import Sum,F
from decimal import Decimal,InvalidOperation
from datetime import datetime, date
from django.utils import timezone



# Create your views here.

def Home(request):
    template = loader.get_template("base.html")
    context = {}
    return HttpResponse(template.render(context, request))

def add_product(request):
    if request.method == 'POST':
        form = ProductForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('productlist')
    else:
        form = ProductForm()
    return render(request, 'add product.html', {'form': form})

def add_production(request):
    if request.method == 'POST':
        form = ProductionForm(request.POST)
        if form.is_valid():
            production = form.save(commit=False)  # Save the form but don't commit yet
            product = production.product  # Get the associated product

            # Debugging output: Before update
            print(f"Before update: Product {product.name} quantity: {product.quantity}")

            # Ensure product quantity is only updated for a new production record
            if not production.pk:  # Check if this is a new record (unsaved)
                product.quantity + production.quantity_added  # Update stock quantity
                product.save()  # Save the product changes

                # Debugging output: After update
                print(f"After update: Product {product.name} quantity: {product.quantity}")
            else:
                print(f"No stock update needed. Production record already exists.")

            production.save()  # Save the production record

            # Debugging output: Production saved
            print(f"Production saved: ID={production.id}, Product={product.name}, Quantity Added={production.quantity_added}")

            # Add production history
            ProductionHistory.objects.create(
                product=product,
                quantity_produced=production.quantity_added,
                date=production.date
            )

            # Debugging output: Production history saved
            print(f"Production history saved for product {product.name} with quantity {production.quantity_added}")

            return redirect('productlist')  # Redirect to product list after successful save
    else:
        form = ProductionForm()

    return render(request, 'addproduction.html', {'form': form})

def product_list(request):
    products = Product.objects.all().order_by('price')  # Sorting by price (ascending)
    return render(request, 'Stock_list.html', {'products': products})

def production_history(request):
    history = ProductionHistory.objects.all()
    return render(request, 'production_history.html', {'history': history})


def get_stock(request, product_id=None):
    try:
        if product_id:
            product = Product.objects.get(id=product_id)
            return JsonResponse({'product_name': product.name, 'stock': product.quantity})
        
        # ✅ Sort products by price (ascending order)
        products = Product.objects.all().order_by('price').values('id', 'name', 'price', 'quantity')
        return JsonResponse({'products': list(products)})

    except Product.DoesNotExist:
        return JsonResponse({'error': 'Product not found'}, status=404)


def createinvoice(request):
    invoice_form = InvoiceForm(request.POST or None)

    if request.method == 'POST':
        if invoice_form.is_valid():
            try:
                with transaction.atomic():
                    # Generate invoice number
                    last_invoice = Invoice.objects.order_by('-id').first()
                    if last_invoice and last_invoice.invoice_number:
                        match = re.search(r'INV2-(\d+)', last_invoice.invoice_number)
                        new_number = int(match.group(1)) + 1 if match else 1
                    else:
                        new_number = 1
                    generated_invoice_number = f"INV2-{new_number:05d}"

                    # Save form data
                    invoice = invoice_form.save(commit=False)
                    invoice.invoice_number = generated_invoice_number
                    invoice.customer = request.POST.get("customer")

                    location_id = request.POST.get("location")
                    customer_obj = None
                    if location_id:
                        try:
                            customer_obj = Customer.objects.get(id=location_id)
                            invoice.location = customer_obj
                        except Customer.DoesNotExist:
                            messages.error(request, "Selected location is invalid.")
                            return redirect('createinvoice')

                    invoice.save()

                    # Get form product and accessory inputs
                    products = request.POST.getlist('products[]')
                    quantities = request.POST.getlist('quantities[]')
                    accessory_quantities = request.POST.getlist('accessory_quantity')
                    accessory_prices = request.POST.getlist('accessory_price')
                    e_way = request.POST.get('e_way', "0")
                    sp_discount = request.POST.get("sp_discount", "0")

                    # Parse decimal inputs
                    try:
                        e_way = Decimal(str(e_way).strip()) if e_way.strip() else Decimal(0)
                    except ValueError:
                        messages.error(request, "Invalid transportation charge entered.")
                        invoice.delete()
                        return redirect('createinvoice')

                    try:
                        sp_discount = Decimal(str(sp_discount).strip()) if sp_discount.strip() else Decimal(0)
                    except InvalidOperation:
                        messages.error(request, "Invalid special discount entered.")
                        invoice.delete()
                        return redirect('createinvoice')

                    if not any(products) and not any(accessory_prices):
                        messages.error(request, "Please add at least one product or accessory.")
                        invoice.delete()
                        return redirect('createinvoice')

                    product_total = Decimal(0)
                    accessory_total = Decimal(0)
                    valid_items = 0

                    # Handle products
                    for i in range(len(products)):
                        try:
                            product_id = products[i].strip()
                            quantity = int(quantities[i].strip()) if quantities[i].strip() else 0

                            if product_id and quantity > 0:
                                product = get_object_or_404(Product, id=product_id)
                                if product.quantity >= quantity:
                                    price = Decimal(str(product.price))
                                    subtotal = price * quantity
                                    product_total += subtotal

                                    InvoiceItem.objects.create(
                                        invoice=invoice,
                                        product=product,
                                        quantity=quantity,
                                        price=price,
                                        subtotal=subtotal
                                    )
                                    product.quantity -= quantity
                                    product.save()
                                    valid_items += 1
                                else:
                                    messages.warning(request, f"Not enough stock for {product.name}. Skipped.")
                        except (ValueError, Product.DoesNotExist) as e:
                            messages.error(request, f"Invalid product selection: {e}")
                            invoice.delete()
                            return redirect('createinvoice')

                    # Handle accessories
                    for j in range(len(accessory_prices)):
                        try:
                            accessory_price = Decimal(str(accessory_prices[j]).strip()) if accessory_prices[j].strip() else Decimal(0)
                            accessory_quantity = int(accessory_quantities[j].strip()) if accessory_quantities[j].strip() else 0

                            if accessory_price > 0 and accessory_quantity > 0:
                                subtotal = accessory_price
                                accessory_total += subtotal

                                InvoiceItem.objects.create(
                                    invoice=invoice,
                                    product=None,
                                    quantity=accessory_quantity,
                                    price=accessory_price,
                                    subtotal=subtotal
                                )
                                valid_items += 1
                        except ValueError as e:
                            messages.error(request, f"Invalid accessory price or quantity entered: {e}")
                            invoice.delete()
                            return redirect('createinvoice')

                    if valid_items == 0:
                        invoice.delete()
                        messages.error(request, "No valid items to create an invoice.")
                        return redirect('createinvoice')

                    # Apply product discount
                    discount_percentage = Decimal(str(invoice.discount_percentage or 0))
                    discount_amount = (product_total * discount_percentage) / Decimal(100)
                    final_product_total = product_total - discount_amount

                    # Calculate full total
                    total_invoice_amount = final_product_total + accessory_total + e_way - sp_discount

                    # Set original total (DO NOT reduce by advance)
                    invoice.total_amount = total_invoice_amount

                    # Deduct advance from money_got only
                    advance_used = Decimal(0)
                    if customer_obj and customer_obj.advance_amount > 0:
                        advance_available = customer_obj.advance_amount
                        if advance_available >= total_invoice_amount:
                            advance_used = total_invoice_amount
                            customer_obj.advance_amount -= total_invoice_amount
                        else:
                            advance_used = advance_available
                            customer_obj.advance_amount = Decimal('0.00')
                        customer_obj.save()

                    invoice.money_got = advance_used
                    invoice.balance_amount = total_invoice_amount - advance_used
                    invoice.e_way = e_way
                    invoice.advance_used = advance_used
                    invoice.save()

                    messages.success(request, f"Invoice {invoice.invoice_number} created successfully!")
                    return redirect('invoicelist')

            except IntegrityError as e:
                messages.error(request, f"Invoice creation failed due to a system error: {e}")
            except Exception as e:
                messages.error(request, f"An unexpected error occurred: {e}")

    products = Product.objects.all().order_by('price')
    location = Customer.objects.all()
    return render(request, 'create invoice.html', {
        'invoice_form': invoice_form,
        'products': products,
        'locations': location,
    })


def invoice_list(request):
    invoices = Invoice.objects.all().order_by('-date','-id')  # Order by date (newest first)
    return render(request, 'invoice list.html', {'invoices': invoices})

def invoice_detail(request, invoice_id):
    invoice = get_object_or_404(Invoice, id=invoice_id)
    items = InvoiceItem.objects.filter(invoice=invoice).order_by('price')
    total_qty = sum(item.quantity for item in items if item.product) 
    subtotal = sum(item.quantity * item.price for item in items)
    discount_amount = (subtotal * invoice.discount_percentage) / 100 if invoice.discount_percentage else 0
    accessory=Invoice.accessory_price
    accessory = Decimal(getattr(invoice, "accessory_price", 0)) 
    e_way=Invoice.e_way
    e_way = Decimal(getattr(invoice, "e_way", 0))
    sp_discount = Invoice.sp_discount
    sp_discount = Decimal(getattr(invoice, "sp_discount", 0)) 
    total_amount = subtotal - discount_amount + accessory + e_way
    context = {
        'invoice': invoice,
        'items': items,
        'subtotal': subtotal,
        'discount_amount': discount_amount,
        'total_amount': total_amount,
        'total_qty': total_qty,
        'accessory':accessory,
        'e_way':e_way,
        'sp_discount':sp_discount,
    }
    return render(request, 'invoice details.html', context)

def Login(request):
    if request.method == 'POST':
        uname = request.POST.get("uname")
        psw = request.POST.get("pswd")
        try:
            l = login.objects.get(uname=uname, pswd=psw)
            if l.utype == "user":
                request.session["uid"] = l.uid_id
                return HttpResponse("<script>alert('Welcome User');window.location='/userhome';</script>")
            elif l.utype == "admin":
                request.session["uid"] = l.uid_id
                return HttpResponse("<script>alert('Welcome Admin');window.location='/adminhome';</script>")

        except login.DoesNotExist:
            return HttpResponse("<script>alert('Invalid username or password.');window.location='/login';</script>")

    template = loader.get_template("login.html")
    context = {}
    return HttpResponse(template.render(context, request))

def registration(request):
    if request.method == "POST":
        r = reg()
        r.name = request.POST.get("Name")
        r.phno = request.POST.get("phno")
        r.save()
        id = reg.objects.latest("id").id
        l = login()
        l.uname = request.POST.get("uname")
        l.pswd = request.POST.get("psw")
        l.utype = 'user'
        l.uid_id = id
        l.save()
        return HttpResponse("<script>alert('Registered successfully');window.location='/login';</script>")

    template = loader.get_template("reg.html")
    context = {}
    return HttpResponse(template.render(context, request))

@never_cache
def userhome(request):
    template = loader.get_template("userhome.html")
    context = {}
    return HttpResponse(template.render(context, request))

def logoutview(request):
    logout(request)  # Logs out the user
    request.session.flush()  # Clears session data

    response = redirect('login')  # Redirects to login page
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'

    return response

def generate_pdf(request, invoice_id):
    # Fetch invoice data from your database
    invoice = get_object_or_404(Invoice, pk=invoice_id)  # Adjust according to your model
    items = invoice.items.all()  # Assuming items are related to the invoice

    # Render the HTML template with the invoice data
    html_content = render_to_string('invoice template.html', {'invoice': invoice, 'items': items})
    
    # Create the response and set the appropriate content type for PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="invoice.pdf"'
    
    # Generate the PDF from the HTML string
    pisa_status = pisa.CreatePDF(html_content, dest=response)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="invoice.pdf"'
    
    # Generate the PDF from the HTML string
    pisa_status = pisa.CreatePDF(html_content, dest=response)
    
    # If there are errors, return an error message
    if pisa_status.err:
        return HttpResponse("Error generating PDF", status=500)

    return response



def delete_invoice(request, invoice_id):
    if request.method == 'POST':  # Only allow POST
        invoice = get_object_or_404(Invoice, id=invoice_id)

        invoice_items = InvoiceItem.objects.filter(invoice=invoice)
        for item in invoice_items:
            if item.product: 
                item.product.quantity += item.quantity
                item.product.save() 

        invoice.delete()  # Now delete the invoice
        messages.success(request, f"Invoice {invoice.invoice_number} deleted successfully! Product stock has been updated.")
        return HttpResponse("<script>alert('Invoice deleted successfully! Product stock has been updated.');window.location='/invoicelist';</script>")

    return HttpResponseNotAllowed(['POST']) 


def adminhome(request):
    template = loader.get_template("adminhome.html")
    context = {}
    return HttpResponse(template.render(context, request))

def sales_report(request):
    invoices = Invoice.objects.all().order_by('-date','-id')
    factory_sales = Factorysale.objects.all()
    today = now().date()
    start_date = request.GET.get('start_date', today)
    end_date = request.GET.get('end_date', today)

    # Apply filtering if both dates are provided
    if start_date and end_date:
        invoices = invoices.filter(date__range=[start_date, end_date])
        factory_sales = factory_sales.filter(date__range=[start_date, end_date])

    # Calculate total sales from both Invoice and Factorysale tables
    total_invoice_sales = sum(invoice.total_amount for invoice in invoices)
    total_factory_sales = sum(factory_sale.total_amount for factory_sale in factory_sales)

    # Overall total sales
    total_sales = total_invoice_sales + total_factory_sales

    context = {
        'invoices': invoices,
        'factory_sales': factory_sales,
        'total_invoice_sales': total_invoice_sales,
        'total_factory_sales': total_factory_sales,
        'total_sales': total_sales,
        'start_date': start_date,
        'end_date': end_date,
    }
    return render(request, 'sales_report.html', context)

def product_sales_report(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if not start_date:
        start_date = datetime.today().date()
    else:
        try:
            start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
        except ValueError:
            return render(request, "product_sales_report.html", {"error": "Invalid start date format"})

    if not end_date:
        end_date = datetime.today().date()
    else:
        try:
            end_date = datetime.strptime(end_date, "%Y-%m-%d").date()
        except ValueError:
            return render(request, "product_sales_report.html", {"error": "Invalid end date format"})
    sales_data = (
        InvoiceItem.objects
        .filter(invoice__date__range=[start_date, end_date]) 
        .values('product__name') 
        .annotate(
            total_quantity=Sum('quantity'), 
        )
        .order_by('-total_quantity')
    )
    return render(request, "product_sales_report.html", {
        "sales_data": sales_data,
        "start_date": start_date,
        "end_date": end_date
    })

def stock_edit(request):
    """View to display and edit stock list by the admin"""
    products = Product.objects.all()

    if request.method == "POST":
        for product in products:
            field_name = f'quantity_{product.id}'
            if field_name in request.POST:
                new_quantity = request.POST[field_name]
                try:
                    new_quantity = int(new_quantity)
                    if new_quantity >= 0:
                        product.quantity = new_quantity
                        product.save()
                    else:
                        messages.error(request, f"Invalid quantity for {product.name}")
                except ValueError:
                    messages.error(request, f"Invalid input for {product.name}")

        messages.success(request, "Stock updated successfully!")
        return redirect('stock_edit')  # Redirect to avoid resubmission

    return render(request, 'stock_edit.html', {'products': products})

import json
from decimal import Decimal

def decimal_to_float(obj):
    if isinstance(obj, Decimal):
        return float(obj)
    raise TypeError


def edit_invoice(request, invoice_id):
    invoice = get_object_or_404(Invoice, id=invoice_id)
    locations = Customer.objects.all()  # Fetch locations, not customers

    if request.method == "POST":
        try:
            # Get POST data
            customer = request.POST.get("customer", invoice.customer)
            location_id = request.POST.get("location")
            date = request.POST.get("date")
            discount_percentage = request.POST.get("discount_percentage", 0)
            accessory_price = request.POST.get("accessory_price", 0)
            e_way = request.POST.get("e_way", 0)
            sp_discount = request.POST.get("sp_discount", 0)
            money_got = request.POST.get("money_got", 0)

            # Convert to Decimal
            discount_percentage = Decimal(discount_percentage or 0)
            accessory_price = Decimal(accessory_price or 0)
            e_way = Decimal(e_way or 0)
            sp_discount = Decimal(sp_discount or 0)
            money_got = Decimal(money_got or 0)

            # Get customer and location objects
            

            if location_id:
                location = Customer.objects.get(id=location_id)  # Correct the location fetch
                invoice.location = location

            # Get product and quantity data
            product_ids = request.POST.getlist("products[]")
            quantities = request.POST.getlist("quantities[]")

            # If no valid products AND all extra fields are zero, return error
            if (not product_ids or not quantities or len(product_ids) != len(quantities)) and accessory_price == 0 and e_way == 0 and sp_discount == 0:
                return JsonResponse({"success": False, "message": "Please add at least one product or accessory/e-way/sp discount."})

            if len(product_ids) != len(quantities):
                return JsonResponse({"success": False, "message": "Mismatched product and quantity count."})

            # Restore stock for existing items in the invoice
            for item in invoice.invoice_items.all():
                if item.product:
                    item.product.quantity += item.quantity
                    item.product.save()
            invoice.invoice_items.all().delete()

            # Process new items
            product_total = Decimal(0)
            skipped_items = []

            for product_id, quantity in zip(product_ids, quantities):
                product = Product.objects.filter(id=product_id).first()

                if not product:
                    skipped_items.append(f"Product ID {product_id} not found.")
                    continue

                try:
                    quantity = int(quantity)
                except ValueError:
                    skipped_items.append(f"Invalid quantity for {product.name}")
                    continue

                if product.quantity < quantity:
                    skipped_items.append(f"Not enough stock for {product.name}")
                    continue

                price = product.price
                subtotal = price * quantity
                product_total += subtotal

                product.quantity -= quantity
                product.save()

                InvoiceItem.objects.create(
                    invoice=invoice,
                    product=product,
                    quantity=quantity,
                    price=price,
                    subtotal=subtotal,
                )

            # Apply discount on product total
            discount_amount = (product_total * discount_percentage) / Decimal(100)
            final_product_total = product_total - discount_amount

            # Final total calculation
            total_amount = final_product_total + accessory_price + e_way - sp_discount
            balance_amount = total_amount - money_got

            # Save invoice
            invoice.date = date
            invoice.discount_percentage = discount_percentage
            invoice.accessory_price = accessory_price
            invoice.e_way = e_way
            invoice.sp_discount = sp_discount
            invoice.money_got = money_got
            invoice.total_amount = total_amount
            invoice.balance_amount = balance_amount
            invoice.save()

            if skipped_items:
                skipped_message = f"Invoice updated, but some items were skipped: {', '.join(skipped_items)}"
                return HttpResponse(f"<script>alert('{skipped_message}');window.location='/invoicelist';</script>")
            else:
                return HttpResponse("<script>alert('Invoice updated successfully!');window.location='/invoicelist';</script>")

        except Exception as e:
            return JsonResponse({"success": False, "message": str(e)})

    # GET method – Load edit form
    invoice_items = InvoiceItem.objects.filter(invoice=invoice, product__isnull=False)
    products = Product.objects.exclude(name__icontains='Accessory')
    customers = Customer.objects.all()

    # Prepare invoice data for the edit form
    invoice_data = {
        "id": invoice.id,
        "customer": str(invoice.customer),
        "location": str(invoice.location),  # Convert location to string for the form
        "date": invoice.date.strftime("%Y-%m-%d"),
        "discount": float(invoice.discount_percentage or 0),
        "accessory_price": float(invoice.accessory_price or 0),
        "e_way": float(invoice.e_way or 0),
        "sp_discount": float(invoice.sp_discount or 0),
        "money_got": float(invoice.money_got or 0),
        "total": float(invoice.total_amount or 0),
        "balance": float(invoice.balance_amount or 0),
        "items": [
            {
                "product_id": item.product.id,
                "quantity": item.quantity,
                "price": float(item.price),
                "subtotal": float(item.subtotal),
            } for item in invoice_items
        ]
    }

    # Render the template
    return render(request, 'Edit invoice.html', {
        'invoice': invoice,
        'invoice_json': json.dumps(invoice_data),
        'products': products,
        'customers': customers,
        'locations': locations  # Pass locations to the template
    })

def create_factory_sale(request):
    if request.method == 'POST':
        form = FactorySaleForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Factory Sale added successfully!")
            return redirect('factorysalelist')  
        else:
            messages.error(request, "Error in form submission.")

    else:
        form = FactorySaleForm()

    return render(request, 'create_factory_sale.html', {'form': form})

def factory_sale_list(request):
    sales = Factorysale.objects.all().order_by('-date')
    return render(request, 'factory_sale_list.html', {'sales': sales})

def delete_factory_sale(request, sale_id):
    sale = get_object_or_404(Factorysale, id=sale_id)
    sale.delete()
    messages.success(request, "Factory Sale deleted successfully!")
    return redirect('factorysalelist')

def addcustomer(request):
    if request.method == 'POST':
        form = CustomerForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('viewcustomer') 
    else:
        form = CustomerForm()
    return render(request, 'Add_customer2.html', {'form': form})

def customer_list(request):
    customers = Customer.objects.all().order_by('name')  # Optional sorting
    total_advance = customers.aggregate(total=Sum('advance_amount'))['total'] or Decimal('0.00')

    return render(request, 'customer_list2.html', {
        'customers': customers,
        'total_advance': total_advance,
    })

def location_report(request):
    location_id = request.GET.get('location')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Set default dates to today's date if not provided
    today = date.today()
    if not start_date:
        start_date = today.strftime('%Y-%m-%d')
    if not end_date:
        end_date = today.strftime('%Y-%m-%d')

    # Base invoice queryset
    invoices = Invoice.objects.all().order_by('date', '-id')

    # Filter by location if selected
    if location_id:
        invoices = invoices.filter(location__id=location_id)

    # Filter by date range
    try:
        parsed_start = datetime.strptime(start_date, "%Y-%m-%d").date()
        parsed_end = datetime.strptime(end_date, "%Y-%m-%d").date()
        invoices = invoices.filter(date__gte=parsed_start, date__lte=parsed_end)
    except ValueError:
        pass  # Optional: handle date parsing errors

    # Sort by date DESC, then ID DESC
    invoices = invoices.order_by('-date', '-id')

    # Calculate totals
    total_sales = invoices.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
    raw_total_balance = invoices.aggregate(total=Sum('balance_amount'))['total'] or Decimal('0')
    if raw_total_balance < 0:
        display_total_balance = f"+{abs(raw_total_balance)}"
    else:
        display_total_balance = str(raw_total_balance)
    total_got = invoices.aggregate(total=Sum('money_got'))['total'] or Decimal('0')

    # Calculate total advance amount for customers in selected location
    customers = Customer.objects.all()
    if location_id:
        customers = customers.filter(id=location_id)

    total_advance = customers.aggregate(total=Sum('advance_amount'))['total'] or 0
    # Handle POST (update money_got)
    if request.method == 'POST':
        count = int(request.POST.get('invoice_count', 0))
        for i in range(1, count + 1):
            invoice_id = request.POST.get(f'invoice_id_{i}')
            money_got = request.POST.get(f'money_got_{i}')
            if invoice_id:
                try:
                    invoice = Invoice.objects.get(id=invoice_id)   
                    added_amount = Decimal(money_got or '0')
                    if added_amount > 0:
                        if invoice.advance_used is None:
                            invoice.advance_used = invoice.money_got
                        invoice.money_got += added_amount
                        invoice.balance_amount = invoice.total_amount - invoice.money_got
                        invoice.save()

                        # Record the payment
                        PaymentRecord.objects.create(
                            invoice=invoice,
                            amount=added_amount,
                            date=timezone.now().date(),
                        )
                except (Invoice.DoesNotExist, ValueError, TypeError) as e:
                    print(f"Error processing invoice {invoice_id}: {e}")
                    continue

        # Redirect to refresh the page with current filters
        return redirect(
            f"{request.path}?location={location_id}&start_date={start_date}&end_date={end_date}"
            if location_id else request.path
        )

    # Add formatted balance for display
    for invoice in invoices:
        if invoice.balance_amount < 0:
            invoice.display_balance_amount = f"+{abs(invoice.balance_amount)}"
        else:
            invoice.display_balance_amount = str(invoice.balance_amount)

    # Render the template with total advance included
    context = {
        'locations': Customer.objects.all(),
        'selected_location_id': int(location_id) if location_id else '',
        'selected_start_date': start_date,
        'selected_end_date': end_date,
        'invoices': invoices,
        'total_sales': total_sales,
        'total_balance': raw_total_balance,
        'display_total_balance': display_total_balance,
        'total_got': total_got,
        'total_advance': total_advance,  # Pass total advance here
    }
    return render(request, 'location_report2.html', context)

import openpyxl
from openpyxl.utils import get_column_letter
from django.utils.dateparse import parse_date

def export_invoices_to_excel(request):
    location_id = request.GET.get('location')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    invoices = Invoice.objects.all()

    # Filter by location
    if location_id:
        invoices = invoices.filter(location__id=location_id)

    # Filter by date range
    if start_date and end_date:
        try:
            parsed_start = parse_date(start_date)
            parsed_end = parse_date(end_date)
            if parsed_start and parsed_end:
                invoices = invoices.filter(date__gte=parsed_start, date__lte=parsed_end)
        except ValueError:
            pass  # Invalid date format

    invoices = invoices.order_by('-date', '-id')

    # Create workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Invoices'

    headers = [
        'Invoice No', 'Date', 'Customer', 'Location',
        'Total Amount', 'Money Got', 'Balance Amount'
    ]
    sheet.append(headers)

    for invoice in invoices:
        sheet.append([
            invoice.invoice_number,
            invoice.date.strftime('%Y-%m-%d'),
            str(invoice.customer),
            str(invoice.location),
            float(invoice.total_amount),
            float(invoice.money_got),
            float(invoice.balance_amount),
        ])

    # Auto-adjust column widths
    for column_cells in sheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 3

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=location_wise_invoices.xlsx'
    workbook.save(response)
    return response


from django.utils import timezone

from django.db.models import Sum


def transaction_report(request):
    # Get the filter parameters from the request
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    selected_location = request.GET.get('location', '')

    # Get distinct locations from Invoice2
    locations = Customer.objects.filter(invoice__isnull=False).distinct()

    # Start with all payments
    payments = PaymentRecord.objects.select_related('invoice')

    # Filter by date
    if start_date:
        payments = payments.filter(date__gte=start_date)
    if end_date:
        payments = payments.filter(date__lte=end_date)

    # Filter by selected location
    if selected_location:
        payments = payments.filter(invoice__location__id=selected_location)

    # Calculate total amount paid
    total_amount = payments.aggregate(Sum('amount'))['amount__sum'] or 0

    context = {
        'payments': payments,
        'start_date': start_date,
        'end_date': end_date,
        'locations': locations,
        'selected_location': locations,
        'total_amount': total_amount,
    }

    return render(request, 'transaction_report2.html', context)

def delete_transaction(request, payment_id):
    # Get the payment record or return 404 if not found
    payment = get_object_or_404(PaymentRecord, pk=payment_id)

    # Delete the payment record
    payment.delete()

    # Get the start_date and end_date from GET parameters to preserve the filter
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Build the redirect URL, including the filter parameters if present
    redirect_url = 'transaction_report'
    if start_date and end_date:
        redirect_url += f"?start_date={start_date}&end_date={end_date}"

    # Redirect to the transaction report page
    return redirect(redirect_url)


def delete_customer(request, customer_id):
    customer = get_object_or_404(Customer, id=customer_id)
    customer.delete()
    return redirect('viewcustomer')

def update_advance_amount(request, customer_id):
    if request.method == 'POST':
        customer = get_object_or_404(Customer, id=customer_id)
        try:
            customer.advance_amount = int(request.POST.get('advance_amount') or 0)
            customer.save()
        except ValueError:
            pass  # Handle invalid input gracefully
    return redirect('viewcustomer') 

def advance_usage_report(request):
    invoices_with_advance = Invoice.objects.filter(advance_used__gt=0)

    total_advance_used = invoices_with_advance.aggregate(total=Sum('advance_used'))['total'] or 0

    return render(request, 'advance_usage_report.html', {
        'invoices': invoices_with_advance,
        'total_advance_used': total_advance_used,
    })